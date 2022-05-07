def init(filename):
    """
    This procedure creates an empty Excel file as nutrition log, with an EMPTY sheet for the personal measures.

    """
    import xlsxwriter as xls
    workbook = xls.Workbook(filename)
    sheet = workbook.add_worksheet('measures')
    headers = ['Last updated', 'Weight', 'Height', 'BMR']
    for row in range(len(headers)):
        sheet.write(row, 0, headers[row])
    workbook.close()
    return True


def get_measures_from_user():
    """
    This procedure gets from the user its personal measures and returns it as a dictionary

    Returns
    -------
    measures [dict]: with key-values of: current weight, height, BMR
    """
    measures = dict()
    measures['weight'] = input("How much you weigh now? [kg]\n")
    measures['height'] = input("What is your height? [cm]\n")
    measures['BMR'] = input("What is your BMR? [kcal]\n")
    # TODO: consider maybe running a weight-log or even circumference log, after mastering pandas
    return measures


def update_worksheet_measures(filename, current_measures, date=None):
    """

    Parameters
    ----------
    current_measures [dict]: user input with key-values of: current weight, height, BMR
    date [str]: The date and time to be documented as last-update status

    Returns
    -------
    This procedure updates the sheet 'measures' with updated measures from the user
    """
    import openpyxl as xls
    from datetime import datetime
    workbook = xls.load_workbook(filename)
    # sheet = workbook.get_sheet_by_name('measures')
    sheet = workbook['measures']
    if date is None:
        date = datetime.now().strftime("%B %d %Y %H:%M")
    values_to_update = [date] + list(current_measures.values())
    row = 1
    col = 'B'
    for value in values_to_update:
        sheet[col+str(row)] = value
        row += 1
    workbook.save(filename)
    workbook.close()
    return True


def read_food_dict():
    """
    The procedure reads the nutritional data of the foods and returns list of food names and values
    Returns
    -------
    food_names [list]: list of food names
    food_dict [dict]:  key = food name and value = {Amount, Units, Calories, Protein, Carbs, Fat}
    """
    import pandas as pd
    # ASSUME: the food dictionary exists, what if not? try-except?
    food_df = pd.read_excel('FoodDictionary.xlsx', index_col=None)
    food_list = food_df.to_dict(orient='records')
    food_names = [value['Food'] for value in food_list]
    food_dict = {row.pop('Food'): row for row in food_list}
    return food_names, food_dict


def init_daily_log(filename, date):
    """
        This procedure creates an empty worksheet related to a specific date, with its related headers.

    """
    import openpyxl as xls
    workbook = xls.load_workbook(filename)
    sheet = workbook.create_sheet(date)
    headers = ['Last Updated', '', 'Balance', 'Total Consumed', 'Total Burned', '']
    for header in headers:
        sheet.append([header])
    sheet['B3'] = 0  # '=B4-B5' # formula for caloric balance DOES NOT WORK TO READ
    sheet['B4'] = 0
    sheet['B5'] = 0
    headers = ['Food', 'Amount', 'Unit', 'Calories', 'Protein', 'Carbs', 'Fats']
    sheet.append(headers)
    workbook.save(filename)
    workbook.close()


def read_daily_log(filename, date):
    """
    The procedure checks if there is logged data for a given date. In case not, it creates an empty sheet for that day.
    Returns
    -------
    food_reads [pandas df]: table of daily food log
    """
    import pandas as pd
    import openpyxl as xls
    from tabulate import tabulate
    workbook = xls.load_workbook(filename)
    # ASSUME: the file itself exists, what if not? try-except?
    if date not in workbook.get_sheet_names():
        init_daily_log(filename, date)
        return None  # or 0
    else:
        workbook.close()
        food_reads = pd.read_excel(filename, sheet_name=date, skiprows=6)
        print(tabulate(food_reads))
    return food_reads


def create_rwo_to_add(food_dict, added_food, amount):
    food_data = food_dict[added_food]
    row_to_add = list(food_data.values())
    factor = amount / (food_data['Amount'])
    row_to_add = [added_food] + ["{:.2f}".format(factor * value) if type(value) is float else value for value in row_to_add]
    return row_to_add


def update_food_log(filename, date, row_to_add, startrow=8):
    from openpyxl import load_workbook
    from datetime import datetime
    import pandas as pd
    workbook = load_workbook(filename)
    sheet = workbook[date]
    cols = ['A', 'B', 'C', 'D', 'E', 'F', 'G']
    for col, value in zip(cols, row_to_add):
        sheet[col+str(startrow)] = value
    sheet['B1'] = datetime.now().strftime("%B %d %Y %H:%M")
    workbook.save(filename)
    workbook.close()
    return True


def sum_daily_log(filename, date, food_reads=None):
    if food_reads is None:
        food_reads = read_daily_log(filename, date)
    headers = ['Protein', 'Carbs', 'Fats']
    total_df = food_reads.sum(axis=0)
    total_values = dict()
    for key in headers:
        total_values[key] = total_df[key]
    total_consumed = total_df['Calories']
    return total_consumed, total_values


def update_total_calories(filename, date, consumed, burned):
    from openpyxl import load_workbook
    from datetime import datetime
    workbook = load_workbook(filename)
    sheet = workbook[date]
    sheet['B1'] = datetime.now().strftime("%B %d %Y %H:%M")
    sheet['B4'] = consumed
    sheet['B5'] = burned
    sheet['B3'] = sheet['B4'].value - sheet['B5'].value
    workbook.save(filename)
    workbook.close()
    return True


def get_total_calories(filename, date):
    from openpyxl import load_workbook
    workbook = load_workbook(filename)  # data_only=True)
    sheet = workbook[date]
    total_cal = dict()
    keys = ['balance', 'consumed', 'burned']
    for index, key in enumerate(keys):
        total_cal[key] = sheet['B'+str(3+index)].value
    updated = sheet['B1'].value
    workbook.close()
    return total_cal, updated


def get_last_update(filename, sheet_name):
    from openpyxl import load_workbook
    workbook = load_workbook(filename)  # data_only=True)
    sheet = workbook[sheet_name]
    updated = sheet['B1'].value
    workbook.close()
    return updated


def init_draft():
    from datetime import datetime
    import pandas as pd
    df_date = pd.DataFrame({'last updated:': datetime.now().strftime("%B %d %Y %H:%M")})
    df_cal_summary = pd.DataFrame({'caloric balance': '0', 'total calories consumed': '0', 'total calories burned:': '0'})
    # ASSUME: I don't need to build the header lines, and will put them as I build the log
    dfs = [df_date, df_cal_summary]
    startrow = 0
    with pd.ExcelWriter('try_'+filename) as writer:
        for df in dfs:
            # FIXME does this updates the sheet? or overrides it?
            df.to_excel(writer, sheet_name=date, engine="xlsxwriter", startrow=startrow)
            startrow += (df.shape[0] + 2)


def get_personal_measures():
    """
    This procedure search if the personal measures of the user exists on the xls file and retrieve it.
    In case the xls does not exist or the data, it gets the personal measures from the user.
    Returns
    -------
    personal_measures {} with keys of: height, BMR, weight
    """
    measures = read_personal_measures()
    if measures is None:
        measures = update_personal_measures()
    return measures


def read_personal_measures():
    """
    This procedure reads the personal measures of the user from the xls file, retrieve it as a dataframe and converts
    it to a dictionary.
    Returns
    -------
    df_measures [dict] with values of: height, BMR, weight
    """
    from os import path
    import pandas as pd
    # ASSUME: if xls exists, than it MUST have the personal measures sheet
    if path.isfile(filename):
        df_measures = pd.read_excel(filename, sheet_name='measures', index_col=None)
        print(df_measures.iloc[:, 1:])
        # index is dropped
        # FIXME consider an alternative way for dealing with the index of the pds-df
        return df_measures.iloc[:, 1:].to_dict(orient='records')[0]
    else:
        return None


def update_personal_measures(filename):
    """
    This procedure gets input from the user personal measures and updates the xls file.
    Returns
    -------
    personal_measures [dict] with keys of: height, BMR, weight
    """
    import pandas as pd
    from os import path
    from openpyxl import load_workbook
    measures = dict()
    measures['weight'] = input("How much you weigh now? [kg]\n")
    measures['height'] = input("What is your height? [cm]\n")
    measures['BMR'] = input("What is your BMR? [kcal]\n")
    df_measures = pd.DataFrame([measures])
    if path.isfile(filename):
        writer = pd.ExcelWriter(filename, engine="openpyxl", mode='a')
        writer.book = load_workbook(filename)
        # notes for the case the xls file was locked for some reason
        # writer.book.security.lockStructure = False
        # writer.close()
    else:
        writer = pd.ExcelWriter(filename, engine="xlsxwriter")
    df_measures.to_excel(writer, sheet_name='measures')
    writer.save()
    return measures


if __name__ == '__main__':
    from os import path
    from datetime import datetime
    filename = 'nutrition_log.xlsx'
    # if not path.isfile(filename):
    #     init()
    #     personal_measures = get_measures_from_user()
    #     update_date = datetime.now().strftime("%B %d %Y %H:%M")
    #     update_worksheet_measures(personal_measures, update_date)
    #
    # food_dict, food_name_list = read_food_dict()
    # update_date = datetime.now().strftime("%B %d %Y %H:%M")
    # date = update_date.split(' 2022')[0]
    # date = 'May 01'
    # init_daily_log(date)
    # startrow = 8
    #
    # added_food = 'Bread'
    # amount = 3
    # food_reads = read_daily_log(date)
    # new_row = create_rwo_to_add(added_food, amount)
# startrow += len(food_reads)
    #
    # consumed = 2200
    # burned = 2900
